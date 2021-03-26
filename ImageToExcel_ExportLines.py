from glob import glob
from os import path
from openpyxl import load_workbook
from itertools import islice
from pathlib import Path
import xlsxwriter
from PIL import Image

# ===> I'm here <===
# Start information 
pathNameExcelRoot = "Label_CMT_20_03_2021_Real.xlsx" # Excel File Root
folderImagesRoot = "31" # Folder Images Root
chunkSize = 1000 # Number Line in Files Export  
# End information

# Get Folder contains Images Root
def TotalImages(folderImagesRoot):
    listNameImages = [path.basename(x).split('.')[0] for x in glob(folderImagesRoot +'/'+ '*.png')]
    listNameImages.sort()
    minListNameImage = int(listNameImages[0])
    maxListNameImage = int(listNameImages[-1])

    return minListNameImage, maxListNameImage

def ReadExcelRoot(pathNameExcelRoot, folderImagesRoot):
    # Get min and max Image
    minListNameImage, maxListNameImage = TotalImages(folderImagesRoot)

    # Read Excel File
    wb = load_workbook(filename=pathNameExcelRoot)

    # Read Current Sheet
    WBsheetData = wb.active

    # Total Row in Excel
    maxRow = (WBsheetData.max_row+1)

    # Start Row in Sheet Root
    start = 1

    # Get Value Cell
    cell_obj = 2
    while start <= maxRow:
        try:
            # Get ID in Row a.k.a Column A
            idValueCellRow = int(WBsheetData.cell(row=cell_obj, column=1).value)
            if idValueCellRow >= minListNameImage:
                startValueCellRow = idValueCellRow
                
                # Get Text in Row a.k.a Column B
                textValueCellRow = (WBsheetData.cell(row=cell_obj, column=2).value)
                DataExport = {}
                while startValueCellRow <= maxListNameImage:
                    idValueCellRow = int(WBsheetData.cell(row=cell_obj, column=1).value)
                    textValueCellRow = (WBsheetData.cell(row=cell_obj, column=2).value)
                    
                    # Set Value to Dictionary
                    DataExport[idValueCellRow]=textValueCellRow

                    start += 1
                    cell_obj += 1
                    startValueCellRow += 1

                if idValueCellRow == maxListNameImage:
                    break
        except Exception as Error:
            print(f"Error in ReadExcelRoot: {Error}")
        start += 1        
        cell_obj += 1
    return DataExport

def ChunkData(pathNameExcelRoot, folderImagesRoot, chunkSize):
    Data = ReadExcelRoot(pathNameExcelRoot, folderImagesRoot)
    it = iter(Data)
    for i in range(0, len(Data), chunkSize):
        yield {k:Data[k] for k in islice(it, chunkSize)}

def ExportToExcels(pathNameExcelRoot, folderImagesRoot, chunkSize):
    # Create folder contain Excels safely
    ExportExcelFolder = "ExportExcelFolder"
    Path(ExportExcelFolder).mkdir(parents=True, exist_ok=True)
    
    for item in ChunkData(pathNameExcelRoot, folderImagesRoot, chunkSize):
        # Name Excel File Output 
        pathNameExcelWrite = next(iter(item))

        # Create an new Excel file and add a worksheet
        workbook = xlsxwriter.Workbook(f"{ExportExcelFolder}/{pathNameExcelWrite}.xlsx")
        worksheet = workbook.add_worksheet()

        # Height Image and Cell
        cell_height = 30.0

        # Set Default Row
        worksheet.set_default_row(int(cell_height))
        worksheet.set_column(1,1, 50)
        numRow = 1
        for k, v in item.items():
            # Insert Text to Rows a.k.a Column A
            worksheet.write_string('A'+str(numRow), str(k))

            # Parameters Image to Row Excel
            cellImage = 'B' + str(numRow)
            filename = f"{folderImagesRoot}/{k}.png"
            img = Image.open(filename)
            image_width, image_height = img.size
            y_scale = cell_height/image_height
            # Insert Image to Row Excel
            worksheet.insert_image(cellImage, filename, {'x_scale': y_scale, 'y_scale': y_scale})

            # Insert Text to Rows a.k.a Column C
            worksheet.write_string('C'+str(numRow), str(v))

            numRow += 1
        # pathNameExcelWrite += 1
        workbook.close()
        print('Ting Ting\n Export:', pathNameExcelWrite)
    print("Mission Success")

# ===> I'm here <===
ExportToExcels(pathNameExcelRoot, folderImagesRoot, chunkSize)
